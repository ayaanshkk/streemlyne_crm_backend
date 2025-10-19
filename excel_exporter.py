import pandas as pd
import os
from datetime import datetime

def export_to_excel(structured_data, customer_name=None):
    """
    Export structured form data to Excel format
    
    Args:
        structured_data (dict): The structured form data
        customer_name (str): Customer name for filename
        
    Returns:
        str: Path to the generated Excel file
    """
    try:
        # Create output directory if it doesn't exist
        output_dir = "generated_excel"
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if customer_name:
            # Clean customer name for filename
            clean_name = "".join(c for c in customer_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            clean_name = clean_name.replace(' ', '_')
            filename = f"{clean_name}_{timestamp}.xlsx"
        else:
            filename = f"bedroom_form_{timestamp}.xlsx"
        
        excel_path = os.path.join(output_dir, filename)
        
        # Create a structured DataFrame
        data_for_excel = []
        
        # Define sections for better organization
        sections = {
            "Customer Information": [
                "customer_name", "address", "room", "tel_mob_number"
            ],
            "Important Dates": [
                "survey_date", "appt_date", "pro_inst_date", "comp_chk_date", "date_deposit_paid"
            ],
            "Style & Colors": [
                "fitting_style", "door_style", "door_colour", "end_panel_colour", 
                "plinth_filler_colour", "worktop_colour", "cabinet_colour", "handles_code_qty_size"
            ],
            "Bedside Cabinets": [
                "bedside_cabinets_floating", "bedside_cabinets_fitted", "bedside_cabinets_freestand", "bedside_cabinets_qty"
            ],
            "Dresser/Desk": [
                "dresser_desk_yes", "dresser_desk_no", "dresser_desk_qty_size"
            ],
            "Internal Mirror": [
                "internal_mirror_yes", "internal_mirror_no", "internal_mirror_qty_size"
            ],
            "Mirror Options": [
                "mirror_silver", "mirror_bronze", "mirror_grey", "mirror_qty"
            ],
            "Soffit Lights": [
                "soffit_lights_spot", "soffit_lights_strip", "soffit_lights_colour",
                "soffit_lights_cool_white", "soffit_lights_warm_white", "soffit_lights_qty"
            ],
            "Gable Lights": [
                "gable_lights_colour", "gable_lights_profile_colour", "gable_lights_black", 
                "gable_lights_white", "gable_lights_qty"
            ],
            "Accessories": [
                "carpet_protection", "floor_tile_protection", "no_floor"
            ],
            "Terms & Signature": [
                "date_terms_conditions_given", "gas_electric_installation_terms_given",
                "customer_signature", "signature_date"
            ]
        }
        
        # Create the data structure
        for section_name, fields in sections.items():
            # Add section header
            data_for_excel.append({
                'Section': section_name,
                'Field': '',
                'Value': '',
                'Notes': ''
            })
            
            # Add fields for this section
            for field in fields:
                field_display = field.replace('_', ' ').title()
                value = structured_data.get(field, '')
                
                # Format value for display
                if value is None:
                    value = ''
                elif isinstance(value, bool):
                    value = '✓' if value else '✗'
                elif value == "✓":
                    value = '✓'
                elif value == "✗":
                    value = '✗'
                else:
                    value = str(value)
                
                data_for_excel.append({
                    'Section': '',
                    'Field': field_display,
                    'Value': value,
                    'Notes': ''
                })
            
            # Add empty row between sections
            data_for_excel.append({
                'Section': '',
                'Field': '',
                'Value': '',
                'Notes': ''
            })
        
        # Create DataFrame
        df = pd.DataFrame(data_for_excel)
        
        # Create Excel writer object
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Bedroom Form Data', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Bedroom Form Data']
            
            # Style the worksheet
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Define styles
            header_font = Font(bold=True, size=12)
            section_font = Font(bold=True, size=11, color='FFFFFF')
            section_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            field_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Style the header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style the data rows
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                section_cell = row[0]
                field_cell = row[1]
                value_cell = row[2]
                notes_cell = row[3]
                
                # Style section headers
                if section_cell.value and field_cell.value == '':
                    for cell in row:
                        cell.font = section_font
                        cell.fill = section_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # Style regular rows
                    for cell in row:
                        cell.font = field_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Special styling for checkmarks
                    if value_cell.value in ['✓', '✗']:
                        value_cell.alignment = Alignment(horizontal='center', vertical='center')
                        value_cell.font = Font(size=12, bold=True)
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 20  # Section
            worksheet.column_dimensions['B'].width = 25  # Field
            worksheet.column_dimensions['C'].width = 20  # Value
            worksheet.column_dimensions['D'].width = 30  # Notes
            
            # Add a summary sheet
            summary_data = []
            if structured_data.get('customer_name'):
                summary_data.append(['Customer Name', structured_data.get('customer_name')])
            if structured_data.get('address'):
                summary_data.append(['Address', structured_data.get('address')])
            if structured_data.get('tel_mob_number'):
                summary_data.append(['Phone', structured_data.get('tel_mob_number')])
            if structured_data.get('survey_date'):
                summary_data.append(['Survey Date', structured_data.get('survey_date')])
            
            # Add processing date
            summary_data.append(['Processed Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            
            summary_df = pd.DataFrame(summary_data, columns=['Field', 'Value'])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Style summary sheet
            summary_sheet = writer.sheets['Summary']
            for cell in summary_sheet[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            
            summary_sheet.column_dimensions['A'].width = 20
            summary_sheet.column_dimensions['B'].width = 30
        
        print(f"Excel file generated successfully: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        raise e